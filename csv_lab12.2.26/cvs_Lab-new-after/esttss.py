from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Tuple, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class PHAAExporter:
    df: pd.DataFrame

    # Form layout
    start_row: int = 12
    rows_per_block: int = 40
    values_per_file: int = 160  # 40 rows * 4 blocks

    def _find_columns(self) -> Tuple[str, str]:
        """Find pH and a/a columns (case-insensitive, tolerant)."""
        def norm(s: str) -> str:
            s = s.strip().lower()
            s = re.sub(r"\s+", "", s)
            s = re.sub(r"[^a-z0-9/]+", "", s)
            return s

        normalized = {col: norm(col) for col in self.df.columns}

        ph_candidates = [c for c, n in normalized.items() if n == "ph"] or \
                        [c for c, n in normalized.items() if "ph" in n]

        aa_candidates = [c for c, n in normalized.items() if n in ("a/a", "aa")] or \
                        [c for c, n in normalized.items() if "a/a" in n or n.endswith("aa")]

        if not ph_candidates:
            raise ValueError("Could not find a pH/PH column.")
        if not aa_candidates:
            raise ValueError("Could not find an a/a column.")

        return ph_candidates[0], aa_candidates[0]

    def _deduplicate(self) -> None:
        """Remove exact duplicate rows."""
        self.df = self.df.drop_duplicates().copy()

    def _ask_missing_aa(self, aa_col: str, ph_col: str) -> None:
        """Ask user to fill missing a/a values."""
        self.df[aa_col] = self.df[aa_col].replace(r"^\s*$", pd.NA, regex=True)
        missing = self.df[aa_col].isna()
        if not missing.any():
            return

        print(f"\nFound {missing.sum()} missing '{aa_col}' values. Fill them now.\n")
        for idx in self.df.index[missing]:
            ph_val = self.df.at[idx, ph_col]
            while True:
                user_in = input(f"Row index {idx} has pH={ph_val!r}. Enter a/a: ").strip()
                if user_in == "":
                    print("  -> a/a cannot be empty.")
                    continue
                self.df.at[idx, aa_col] = int(user_in) if re.fullmatch(r"\d+", user_in) else user_in
                break

    def _extract_pairs(self, aa_col: str, ph_col: str) -> List[Tuple[int, float]]:
        """
        Return cleaned list of (a/a, pH) where:
        - a/a is a real integer (no 41.0)
        - pH is a float
        """
        small = self.df[[aa_col, ph_col]].copy()

        # Drop missing
        small = small.dropna(subset=[aa_col, ph_col])

        # Convert pH to numeric
        small[ph_col] = pd.to_numeric(small[ph_col], errors="coerce")
        small = small.dropna(subset=[ph_col])

        # Convert a/a to numeric (NaNs happen if there were blanks or non-numeric)
        small[aa_col] = pd.to_numeric(small[aa_col], errors="coerce")
        small = small.dropna(subset=[aa_col])

        # Force real integer dtype (this removes .0)
        small[aa_col] = small[aa_col].astype("int64")

        # Build pairs as (int, float)
        return [(int(aa), float(ph)) for aa, ph in zip(small[aa_col], small[ph_col])]

    def _write_into_template(self, ws: Worksheet, pairs: List[Tuple[str, float]]) -> None:
        """
        Fill the template form like your screenshot:
        A,C,E,G -> a/a
        B,D,F,H -> pH
        Rows start at 12, 40 rows per block.
        """
        aa_cols = ("A", "C", "E", "G")
        ph_cols = ("B", "D", "F", "H")

        max_items = min(len(pairs), self.values_per_file)

        for i in range(max_items):
            block = i // self.rows_per_block      # 0..3
            offset = i % self.rows_per_block      # 0..39
            row = self.start_row + offset

            aa, ph = pairs[i]
            ws[f"{aa_cols[block]}{row}"] = int(aa)
            ws[f"{ph_cols[block]}{row}"] = float(ph)

    def export_to_multiple_files(self, template_path: str = "pH.xlsx", out_prefix: str = "ph_") -> List[str]:
        """
        Opens template_path, writes first 160 pairs, saves as ph_1.xlsx.
        Next 160 -> ph_2.xlsx, etc.
        """
        ph_col, aa_col = self._find_columns()
        self._deduplicate()
        self._ask_missing_aa(aa_col=aa_col, ph_col=ph_col)

        pairs = self._extract_pairs(aa_col=aa_col, ph_col=ph_col)
        if not pairs:
            raise ValueError("No valid (a/a, pH) pairs found after cleaning.")

        total_files = (len(pairs) + self.values_per_file - 1) // self.values_per_file
        saved_files: List[str] = []

        for k in range(total_files):
            start = k * self.values_per_file
            end = start + self.values_per_file
            chunk = pairs[start:end]

            # 1) Load the template fresh each time (clean form)
            wb = load_workbook(template_path)
            ws = wb.active  # or wb["ENTYPO KATAGRAFHS pH"] if you need specific sheet name

            # 2) Fill it
            self._write_into_template(ws, chunk)

            # 3) Save as ph_1.xlsx, ph_2.xlsx, ...
            out_path = f"{out_prefix}{k+1}.xlsx"
            wb.save(out_path)
            saved_files.append(out_path)

        return saved_files


# Example usage:
df = pd.read_excel("16052024-6.xlsx")
exporter = PHAAExporter(df)
files = exporter.export_to_multiple_files(template_path="pH.xlsx", out_prefix="ph_")
print("Saved:", files)
