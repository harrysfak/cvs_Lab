import openpyxl

template = "pH.xlsx"
columns_to_interact = ["B", "D", "F", "H"]


wb = openpyxl.load_workbook(template)
ws = wb.active

try: # runs for every cell value:


    # next need to make conditions for 
    for i in range(10,100):
        ws[f"B{i}"].value = i
        print(ws[f"B{i}"].value)

except ValueError as e:
    print(e)
    print(i)


wb.save(f'balances{i}.xlsx')