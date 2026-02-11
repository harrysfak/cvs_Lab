"""
This module collects pH values connected with their a/a value and apply them into the structured form xlsx.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

import config


class PHHandler:
    def __init__(self, df: pd.DataFrame, aa_col: str = "a/a", ph_col: str = "pH"):
        self.df = df.copy()
        self.aa_col = aa_col
        self.ph_col = ph_col
        self.df.columns = self.df.columns.astype(str).str.strip()
        self.df = self.df.loc[:, ~self.df.columns.duplicated(keep="first")]

        # normalize column names for pH (PH/ph/pH -> pH)
        self._normalize_columns()

        self.pairs_df = self._extract_pairs()


    def _normalize_columns(self):
        cols = {c.strip(): c for c in self.df.columns}  # map stripped -> original

        # normalize a/a if needed (optional: keep yours simple)
        # if self.aa_col not in self.df.columns and "a/a" in cols:
        #     self.df = self.df.rename(columns={cols["a/a"]: self.aa_col})

        # normalize pH variants to self.ph_col
        for candidate in ["pH", "PH", "ph", "Ph", "pH "]:
            if candidate in cols and self.ph_col not in self.df.columns:
                self.df = self.df.rename(columns={cols[candidate]: self.ph_col})
                break

    def _extract_pairs(self) -> pd.DataFrame:
        if self.aa_col not in self.df.columns or self.ph_col not in self.df.columns:
            raise KeyError(f"Λείπει στήλη: '{self.aa_col}' ή '{self.ph_col}'")

        pairs = self.df[[self.aa_col, self.ph_col]].copy()
        print("1aa_col:", self.aa_col, type(self.aa_col))
        print("ph_col:", self.ph_col, type(self.ph_col))
        pairs[self.aa_col] = pd.to_numeric(pairs[self.aa_col], errors="coerce")
        pairs[self.ph_col] = pd.to_numeric(pairs[self.ph_col], errors="coerce")
        print("2aa_col:", self.aa_col, type(self.aa_col))
        print("ph_col:", self.ph_col, type(self.ph_col))
        # κρατάμε μόνο rows που έχουν aa
        pairs = pairs.dropna(subset=[self.aa_col])

        # μοναδικό aa (κρατάμε last)
        pairs = pairs.drop_duplicates(subset=[self.aa_col], keep="last")

        pairs[self.aa_col] = pairs[self.aa_col].astype(int)

        return pairs.sort_values(self.aa_col).reset_index(drop=True)

    def to_map(self) -> dict[int, float]:
        ok = self.pairs_df.dropna(subset=[self.ph_col])
        return dict(zip(ok[self.aa_col].astype(int), ok[self.ph_col].astype(float)))

    def missing_ph_aas(self) -> list[int]:
        missing = self.pairs_df[self.pairs_df[self.ph_col].isna()][self.aa_col]
        return missing.astype(int).tolist()

    # ---------- FORM WRITING ----------

    @staticmethod
    def _slot_to_form_cell(slot: int, start_row: int, block_size: int,
                           start_col_idx: int, col_step: int) -> str:
        """
        slot=1..150 -> A3..A52, D3..D52, G3..G52
        (δηλαδή 3 blocks των 50)
        """
        block = (slot - 1) // block_size
        pos = (slot - 1) % block_size
        row = start_row + pos
        col_idx = start_col_idx + block * col_step
        return f"{get_column_letter(col_idx)}{row}"

    def fill_form(
        self,
        template_path: str,
        out_path: str,
        sheet_name: str | None = None,
        start_row: int = 3,
        block_size: int = 50,
        max_per_form: int = 150,     # <-- ΝΕΟ (50*3)
        ph_start_col: int = 1,       # A
        col_step: int = 3,           # A, D, G
        write_aa: bool = True,
        aa_col_offset: int = 1,      # δίπλα δεξιά: A->B, D->E, G->H
        strict_missing_ph: bool = True
    ) -> list[str]:
        """
        Αν τα δείγματα > max_per_form, δημιουργεί πολλαπλά αρχεία φόρμας
        και κάθε νέα φόρμα ξεκινά πάλι από A3.

        Επιστρέφει λίστα με paths των αρχείων που δημιουργήθηκαν.
        """
        missing = self.missing_ph_aas()
        if missing and strict_missing_ph:
            raise ValueError(
                f"Λείπει pH για a/a: {missing[:20]}" + (" ..." if len(missing) > 20 else "")
            )

        # παίρνουμε ζευγάρια (aa, pH) ταξινομημένα
        pairs = self.pairs_df.copy()
        pairs = pairs.sort_values(self.aa_col).reset_index(drop=True)

        # αν δεν είσαι strict, απλά πετάμε όσα δεν έχουν pH
        pairs = pairs.dropna(subset=[self.ph_col])

        records = list(zip(pairs[self.aa_col].astype(int), pairs[self.ph_col].astype(float)))
        if not records:
            return []

        # split σε chunks των max_per_form
        chunks = [records[i:i + max_per_form] for i in range(0, len(records), max_per_form)]

        base, ext = os.path.splitext(out_path)
        out_paths: list[str] = []

        for part_idx, chunk in enumerate(chunks, start=1):
            # φτιάχνουμε νέο output path για κάθε φόρμα
            part_path = out_path if len(chunks) == 1 else f"{base}_part{part_idx}{ext}"

            wb = load_workbook(template_path)
            ws = wb[sheet_name] if sheet_name else wb.active

            # γράφουμε chunk με slot index που ξεκινά από 1
            for slot, (aa, ph) in enumerate(chunk, start=1):
                ph_cell = self._slot_to_form_cell(
                    slot=slot,
                    start_row=start_row,
                    block_size=block_size,
                    start_col_idx=ph_start_col,
                    col_step=col_step
                )
                ws[ph_cell].value = round(ph, 2)

                if write_aa:
                    col_letters = "".join(ch for ch in ph_cell if ch.isalpha())
                    row_numbers = int("".join(ch for ch in ph_cell if ch.isdigit()))
                    ph_col_idx = column_index_from_string(col_letters)
                    aa_cell = f"{get_column_letter(ph_col_idx + aa_col_offset)}{row_numbers}"
                    ws[aa_cell].value = aa

            wb.save(part_path)
            out_paths.append(part_path)

        return out_paths
